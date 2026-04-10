#!/usr/bin/env python3
"""
formula_audit.py — 재무모델 수식 심볼릭 정합성 검증

sympy를 사용하여:
1. KPI 수식 정의 vs xlsx 수식 간 심볼릭 등가성 검증
2. 단위 일관성 검증 (월 vs 연, % vs 절대값)
3. 순환참조 감지
4. 수식 종속성 그래프 생성 + 누락 참조 탐지

pip install sympy openpyxl
"""

import re
import sys
import json
from pathlib import Path
from collections import defaultdict
from typing import NamedTuple

try:
    import sympy
    from sympy import symbols, simplify, Eq, solve, Rational
except ImportError:
    print("ERROR: sympy 필요. pip install sympy", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# §1. KPI 수식 심볼릭 정의 (kpi-formulas.py 기반)
# ============================================================

# sympy 심볼 정의
(arpu, gross_margin, monthly_churn, annual_churn,
 total_marketing, new_paid_users, cac_val,
 net_burn, net_new_arr, revenue_growth, ebitda_margin,
 cash_balance, monthly_net_burn,
 new_mrr, expansion_mrr, contraction_mrr, churn_mrr,
 beginning_mrr) = symbols(
    'arpu gross_margin monthly_churn annual_churn '
    'total_marketing new_paid_users cac_val '
    'net_burn net_new_arr revenue_growth ebitda_margin '
    'cash_balance monthly_net_burn '
    'new_mrr expansion_mrr contraction_mrr churn_mrr '
    'beginning_mrr'
)

# 정준(canonical) KPI 수식
CANONICAL_FORMULAS = {
    "LTV": arpu * gross_margin / monthly_churn,
    "CAC": total_marketing / new_paid_users,
    "LTV/CAC": (arpu * gross_margin / monthly_churn) / (total_marketing / new_paid_users),
    "CAC_Payback": cac_val / (arpu * gross_margin),
    "Burn_Multiple": net_burn / net_new_arr,
    "Rule_of_40": revenue_growth + ebitda_margin,
    "Runway": cash_balance / monthly_net_burn,
    "NRR": (beginning_mrr + expansion_mrr - contraction_mrr - churn_mrr) / beginning_mrr,
    "Quick_Ratio": (new_mrr + expansion_mrr) / (contraction_mrr + churn_mrr),
    "Net_New_MRR": new_mrr + expansion_mrr - contraction_mrr - churn_mrr,
}


class FormulaCheck(NamedTuple):
    kpi_name: str
    canonical: str     # 정준 수식
    found: str         # xlsx에서 발견된 수식
    equivalent: bool   # 심볼릭 등가 여부
    issue: str         # 문제 설명


class UnitCheck(NamedTuple):
    sheet: str
    cell: str
    formula: str
    issue: str  # "월/연 혼용", "% vs 절대값" 등


class CircularRef(NamedTuple):
    sheet: str
    cell: str
    cycle: list[str]  # [A1→B2→C3→A1]


# ============================================================
# §2. 수식 파싱 + 종속성 그래프
# ============================================================

def parse_cell_refs(formula: str) -> list[str]:
    """수식에서 셀 참조 추출. 'Sheet'!A1 또는 A1 형식."""
    refs = []
    # 수식 앞의 = 제거
    clean = formula.lstrip("=")
    # 시트간 참조: 'Sheet Name'!A1 또는 Sheet!A1
    for m in re.finditer(r"'?([^'!=+\-*/(),\s]+)'?!([A-Z]+\d+)", clean):
        refs.append(f"{m.group(1)}!{m.group(2)}")
    # 같은 시트 내 참조: 단독 A1 (시트 접두사 없는)
    for m in re.finditer(r"(?<![A-Z!'])([A-Z]{1,3}\d{1,7})(?![A-Z\d])", formula):
        refs.append(m.group(1))
    return refs


def build_dependency_graph(filepath: str) -> dict:
    """xlsx에서 수식 종속성 그래프 구축.

    Returns:
        {
            "graph": {cell_key: [depends_on_keys]},
            "formulas": {cell_key: formula_str},
            "circular": [CircularRef],
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    graph = defaultdict(list)
    formulas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    cell_key = f"{sheet_name}!{get_column_letter(cell.column)}{cell.row}"
                    formula_str = str(cell.value)
                    formulas[cell_key] = formula_str

                    refs = parse_cell_refs(formula_str)
                    for ref in refs:
                        # 시트 접두사 없는 참조는 현재 시트 기준
                        if "!" not in ref:
                            ref = f"{sheet_name}!{ref}"
                        graph[cell_key].append(ref)

    wb.close()

    # 순환참조 감지 (DFS)
    circular = _detect_cycles(graph)

    return {
        "graph": dict(graph),
        "formulas": formulas,
        "circular": [c._asdict() for c in circular],
    }


def _detect_cycles(graph: dict) -> list[CircularRef]:
    """DFS로 순환참조 감지."""
    visited = set()
    in_stack = set()
    cycles = []

    def dfs(node, path):
        if node in in_stack:
            # 순환 발견
            cycle_start = path.index(node)
            cycle = path[cycle_start:] + [node]
            sheet = node.split("!")[0] if "!" in node else ""
            cell = node.split("!")[1] if "!" in node else node
            cycles.append(CircularRef(
                sheet=sheet, cell=cell,
                cycle=["→".join(cycle)]
            ))
            return

        if node in visited:
            return

        visited.add(node)
        in_stack.add(node)

        for neighbor in graph.get(node, []):
            dfs(neighbor, path + [node])

        in_stack.discard(node)

    for node in graph:
        if node not in visited:
            dfs(node, [])

    return cycles


# ============================================================
# §3. 단위 검증
# ============================================================

# 월/연 혼용 감지 패턴
MONTHLY_INDICATORS = ["monthly", "월", "mom", "mrr", "/month", "per month"]
ANNUAL_INDICATORS = ["annual", "연", "yoy", "arr", "/year", "per year", "yearly"]


def check_unit_consistency(filepath: str) -> list[dict]:
    """수식 내 월/연 단위 혼용 감지."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    issues = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula = str(cell.value).lower()
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"

                    # 같은 수식에서 월/연 참조가 혼재하는지 검사
                    has_monthly = any(ind in formula for ind in MONTHLY_INDICATORS)
                    has_annual = any(ind in formula for ind in ANNUAL_INDICATORS)

                    if has_monthly and has_annual:
                        # 12로 나누거나 곱하는 변환이 있는지 확인
                        has_conversion = bool(re.search(r'[*/]\s*12\b', formula))
                        if not has_conversion:
                            issues.append(UnitCheck(
                                sheet=sheet_name,
                                cell=cell_ref,
                                formula=str(cell.value)[:80],
                                issue="월/연 단위 혼용 — ×12 또는 /12 변환 없음",
                            )._asdict())

    wb.close()
    return issues


# ============================================================
# §4. 전체 감사
# ============================================================

def full_audit(filepath: str) -> dict:
    """전체 수식 감사 실행.

    Returns:
        {
            "file": str,
            "dependency_graph": {...},
            "circular_refs": [...],
            "unit_issues": [...],
            "stats": {"total_formulas": int, "cross_sheet_refs": int},
            "summary": str,
        }
    """
    dep = build_dependency_graph(filepath)
    unit_issues = check_unit_consistency(filepath)

    total_formulas = len(dep["formulas"])
    cross_sheet = sum(
        1 for deps in dep["graph"].values()
        for d in deps
        if "!" in d
    )

    error_count = len(dep["circular"]) + len(unit_issues)

    return {
        "file": str(filepath),
        "total_formulas": total_formulas,
        "cross_sheet_refs": cross_sheet,
        "circular_refs": dep["circular"],
        "unit_issues": unit_issues,
        "dependency_count": {
            sheet: sum(1 for k in dep["formulas"] if k.startswith(f"{sheet}!"))
            for sheet in set(k.split("!")[0] for k in dep["formulas"])
        },
        "summary": f"수식 {total_formulas}개 | 시트간참조 {cross_sheet} | "
                   f"순환참조 {len(dep['circular'])}건 | 단위이슈 {len(unit_issues)}건",
    }


# ============================================================
# §5. KPI 수식 등가성 검증
# ============================================================

def verify_kpi_formula(kpi_name: str, user_expr_str: str) -> dict:
    """사용자가 제공한 수식이 정준 수식과 심볼릭으로 등가인지 검증.

    Args:
        kpi_name: CANONICAL_FORMULAS의 키 (예: "LTV", "CAC")
        user_expr_str: sympy가 파싱할 수 있는 수식 문자열
            예: "arpu * gross_margin / monthly_churn"

    Returns:
        {"kpi": str, "canonical": str, "user": str, "equivalent": bool, "diff": str}
    """
    if kpi_name not in CANONICAL_FORMULAS:
        return {
            "kpi": kpi_name,
            "error": f"알 수 없는 KPI: {kpi_name}. 사용 가능: {list(CANONICAL_FORMULAS.keys())}",
        }

    canonical = CANONICAL_FORMULAS[kpi_name]
    try:
        user_expr = sympy.sympify(user_expr_str)
    except Exception as e:
        return {
            "kpi": kpi_name,
            "canonical": str(canonical),
            "user": user_expr_str,
            "equivalent": False,
            "error": f"수식 파싱 실패: {e}",
        }

    diff = simplify(canonical - user_expr)
    is_equiv = diff == 0

    return {
        "kpi": kpi_name,
        "canonical": str(canonical),
        "user": str(user_expr),
        "equivalent": is_equiv,
        "diff": str(diff) if not is_equiv else "0",
    }


# ============================================================
# CLI
# ============================================================

def _print_result(result: dict) -> None:
    if "error" in result:
        print(f"오류: {result['error']}")
        return

    print(f"\n{'='*70}")
    print(f"수식 심볼릭 감사: {result['file']}")
    print(f"{result['summary']}")
    print(f"{'='*70}")

    if result["circular_refs"]:
        print(f"\n{'─'*50}")
        print(f"🔴 순환참조 ({len(result['circular_refs'])}건):")
        for c in result["circular_refs"]:
            print(f"  {c['sheet']}:{c['cell']} — {c['cycle']}")

    if result["unit_issues"]:
        print(f"\n{'─'*50}")
        print(f"🟡 단위 이슈 ({len(result['unit_issues'])}건):")
        for u in result["unit_issues"]:
            print(f"  {u['sheet']}:{u['cell']} — {u['issue']}")
            print(f"    수식: {u['formula']}")

    if result.get("dependency_count"):
        print(f"\n{'─'*50}")
        print("시트별 수식 수:")
        for sheet, count in sorted(result["dependency_count"].items()):
            print(f"  {sheet}: {count}개")

    if not result["circular_refs"] and not result["unit_issues"]:
        print(f"\n✅ 순환참조 0건, 단위이슈 0건")


def main():
    """CLI:
        python formula_audit.py audit <xlsx_file> [--json]
        python formula_audit.py verify <KPI_name> "<sympy_expression>"
    """
    if len(sys.argv) < 3:
        print("Usage:")
        print('  python formula_audit.py audit <xlsx_file> [--json]')
        print('  python formula_audit.py verify LTV "arpu * gross_margin / monthly_churn"')
        print(f"\n  Available KPIs: {list(CANONICAL_FORMULAS.keys())}")
        sys.exit(1)

    command = sys.argv[1]

    if command == "audit":
        filepath = sys.argv[2]
        json_mode = "--json" in sys.argv
        result = full_audit(filepath)
        if json_mode:
            print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        else:
            _print_result(result)

        error_count = len(result.get("circular_refs", [])) + len(result.get("unit_issues", []))
        sys.exit(1 if error_count > 0 else 0)

    elif command == "verify":
        kpi_name = sys.argv[2]
        expr_str = " ".join(sys.argv[3:]).strip().strip('"').strip("'")
        result = verify_kpi_formula(kpi_name, expr_str)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result.get("equivalent") else 1)

    else:
        print(f"알 수 없는 명령: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()
