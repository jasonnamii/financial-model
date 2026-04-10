#!/usr/bin/env python3
"""
recalc.py — xlsx 재무모델 재계산 검증기

openpyxl로 xlsx를 로드하여:
1. 수식 셀 전수 탐지 + 매핑
2. 탭간 참조(시트간링크) 정합성 검증
3. 하드코딩 감지 (Assumptions 탭 외 숫자 직입력)
4. 색상규약 검증 (파랑=입력, 검정=수식, 초록=시트간링크)
5. 시나리오 스위치 셀 존재 확인
6. 단방향 흐름 검증 (Assumptions→Calc→Output, 역참조 감지)

pip install openpyxl
"""

import re
import sys
import json
from pathlib import Path
from typing import NamedTuple
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 탭 분류 (단방향 흐름 검증용)
# ============================================================

TAB_TIERS = {
    "input": ["assumptions", "input", "inputs", "가정"],
    "calc": ["revenue", "costs", "cost", "매출", "비용", "calculation"],
    "output": ["p&l", "pl", "pnl", "cash flow", "cashflow", "kpis", "kpi",
               "scenarios", "scenario", "손익", "현금흐름", "시나리오", "대시보드"],
}

TIER_ORDER = {"input": 0, "calc": 1, "output": 2, "unknown": 3}

# 색상규약 (RGB hex)
COLOR_INPUT = "0000FF"     # 파랑 = 입력
COLOR_FORMULA = "000000"   # 검정 = 수식
COLOR_CROSS_REF = "008000" # 초록 = 시트간링크
# 허용 범위 확장 (근사 색상 포함)
BLUE_FAMILY = {"0000FF", "0070C0", "4472C4", "0000CC", "1F4E79"}
GREEN_FAMILY = {"008000", "00B050", "00B0F0", "70AD47", "548235"}
BLACK_FAMILY = {"000000", "333333", "404040", "262626"}


class CellInfo(NamedTuple):
    sheet: str
    cell: str
    value: object
    is_formula: bool
    formula: str
    font_color: str
    refs_other_sheets: list  # [sheet_name, ...]


class AuditError(NamedTuple):
    category: str  # hardcode | color | flow | cross_ref | scenario
    severity: str  # error | warning
    sheet: str
    cell: str
    message: str


def classify_tab(sheet_name: str) -> str:
    """시트명으로 tier 분류."""
    name_lower = sheet_name.lower().strip()
    for tier, keywords in TAB_TIERS.items():
        if any(kw in name_lower for kw in keywords):
            return tier
    return "unknown"


def extract_sheet_refs(formula: str, all_sheets: list[str]) -> list[str]:
    """수식에서 참조하는 다른 시트명 추출."""
    refs = set()
    # 수식 앞의 = 제거
    clean = formula.lstrip("=")
    # 패턴: 'Sheet Name'!A1 또는 SheetName!A1
    for m in re.finditer(r"'?([^'!=+\-*/(),\s]+)'?!", clean):
        ref_sheet = m.group(1)
        if ref_sheet in all_sheets:
            refs.add(ref_sheet)
    return list(refs)


def get_font_color_hex(cell) -> str:
    """셀의 폰트 색상을 hex로 반환."""
    try:
        color = cell.font.color
        if color is None:
            return "000000"
        if color.rgb and color.rgb != "00000000":
            # openpyxl은 ARGB 형식 (8자리) — 앞 2자리 alpha 제거
            rgb = str(color.rgb)
            if len(rgb) == 8:
                return rgb[2:]
            return rgb
        if color.theme is not None:
            return f"theme:{color.theme}"
        return "000000"
    except Exception:
        return "000000"


def audit_workbook(filepath: str) -> dict:
    """xlsx 재무모델 전수 감사.

    Returns:
        {
            "file": str,
            "sheets": [{"name": str, "tier": str, "rows": int, "formulas": int, "values": int}],
            "errors": [AuditError as dict],
            "stats": {"total_cells": int, "formula_cells": int, "value_cells": int, "cross_refs": int},
            "summary": str,
        }
    """
    path = Path(filepath)
    if not path.exists():
        return {"error": f"파일 없음: {filepath}"}

    wb = openpyxl.load_workbook(filepath, data_only=False)
    all_sheets = wb.sheetnames
    errors: list[AuditError] = []
    sheet_infos = []

    total_cells = 0
    formula_cells = 0
    value_cells = 0
    cross_ref_count = 0

    # 시나리오 스위치 존재 여부
    scenario_switch_found = False

    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        tier = classify_tab(sheet_name)
        sheet_formulas = 0
        sheet_values = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                total_cells += 1
                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                is_formula = str(cell.value).startswith("=")
                formula_str = str(cell.value) if is_formula else ""
                font_color = get_font_color_hex(cell)

                if is_formula:
                    formula_cells += 1
                    sheet_formulas += 1

                    # 시트간 참조 추출
                    sheet_refs = extract_sheet_refs(formula_str, all_sheets)
                    if sheet_refs:
                        cross_ref_count += len(sheet_refs)

                        # 단방향 흐름 검증 — output/calc이 input을 역참조하는 것은 OK,
                        # input이 calc/output을 참조하면 역흐름
                        current_tier = tier
                        for ref_sheet in sheet_refs:
                            ref_tier = classify_tab(ref_sheet)
                            if (TIER_ORDER.get(current_tier, 99) <
                                    TIER_ORDER.get(ref_tier, 99)):
                                errors.append(AuditError(
                                    category="flow",
                                    severity="error",
                                    sheet=sheet_name,
                                    cell=cell_ref,
                                    message=f"역흐름 감지: {tier}({sheet_name}) → {ref_tier}({ref_sheet}). "
                                            f"단방향 원칙 위반(Assumptions→Calc→Output)",
                                ))

                        # 색상규약: 시트간 참조 = 초록
                        if font_color not in GREEN_FAMILY and not font_color.startswith("theme:"):
                            errors.append(AuditError(
                                category="color",
                                severity="warning",
                                sheet=sheet_name,
                                cell=cell_ref,
                                message=f"시트간링크 색상 위반: 초록 필요, 실제 #{font_color}",
                            ))

                    else:
                        # 일반 수식 = 검정
                        if font_color not in BLACK_FAMILY and not font_color.startswith("theme:"):
                            errors.append(AuditError(
                                category="color",
                                severity="warning",
                                sheet=sheet_name,
                                cell=cell_ref,
                                message=f"수식 색상 위반: 검정 필요, 실제 #{font_color}",
                            ))

                elif isinstance(cell.value, (int, float)):
                    value_cells += 1
                    sheet_values += 1

                    # 하드코딩 감지: Assumptions 탭 외에서 숫자 직입력
                    if tier != "input":
                        # 예외: 0, 1, 12, 100 등 상수는 허용
                        if cell.value not in (0, 1, -1, 12, 100, 365, 52, 4):
                            errors.append(AuditError(
                                category="hardcode",
                                severity="error",
                                sheet=sheet_name,
                                cell=cell_ref,
                                message=f"하드코딩 감지: 값={cell.value}. "
                                        f"Assumptions 탭에서 참조해야 함",
                            ))

                    else:
                        # Input 탭의 값 = 파랑
                        if font_color not in BLUE_FAMILY and not font_color.startswith("theme:"):
                            errors.append(AuditError(
                                category="color",
                                severity="warning",
                                sheet=sheet_name,
                                cell=cell_ref,
                                message=f"입력값 색상 위반: 파랑 필요, 실제 #{font_color}",
                            ))

                # 시나리오 스위치 탐지
                if isinstance(cell.value, str) and any(
                    kw in cell.value.lower()
                    for kw in ["scenario", "시나리오", "switch", "스위치", "base", "bull", "bear"]
                ):
                    scenario_switch_found = True

        sheet_infos.append({
            "name": sheet_name,
            "tier": tier,
            "rows": ws.max_row or 0,
            "formulas": sheet_formulas,
            "values": sheet_values,
        })

    # 시나리오 스위치 미발견
    if not scenario_switch_found:
        errors.append(AuditError(
            category="scenario",
            severity="warning",
            sheet="(전체)",
            cell="-",
            message="시나리오 스위치 셀 미발견. Assumptions 탭에 Base/Bull/Bear 스위치 필요",
        ))

    # 에러/경고 집계
    error_count = sum(1 for e in errors if e.severity == "error")
    warning_count = sum(1 for e in errors if e.severity == "warning")

    wb.close()

    return {
        "file": str(filepath),
        "sheets": sheet_infos,
        "errors": [e._asdict() for e in errors],
        "stats": {
            "total_cells": total_cells,
            "formula_cells": formula_cells,
            "value_cells": value_cells,
            "cross_refs": cross_ref_count,
        },
        "summary": f"셀 {total_cells}개 검사 | 수식 {formula_cells} | 값 {value_cells} | "
                   f"시트간참조 {cross_ref_count} | 에러 {error_count} | 경고 {warning_count}",
    }


# ============================================================
# CLI
# ============================================================

def _print_result(result: dict) -> None:
    if "error" in result:
        print(f"오류: {result['error']}")
        return

    print(f"\n{'='*70}")
    print(f"재무모델 재계산 검증: {result['file']}")
    print(f"{result['summary']}")
    print(f"{'='*70}")

    # 시트 요약
    print(f"\n{'─'*50}")
    print("시트 구성:")
    for s in result["sheets"]:
        print(f"  [{s['tier']:>7}] {s['name']:<25} "
              f"행:{s['rows']:<5} 수식:{s['formulas']:<5} 값:{s['values']}")

    # 에러 (error 먼저, warning 나중)
    errors = result["errors"]
    if errors:
        by_severity = defaultdict(list)
        for e in errors:
            by_severity[e["severity"]].append(e)

        for sev in ["error", "warning"]:
            items = by_severity.get(sev, [])
            if items:
                label = "🔴 ERROR" if sev == "error" else "🟡 WARNING"
                print(f"\n{'─'*50}")
                print(f"{label} ({len(items)}건):")
                # 카테고리별 그룹핑
                by_cat = defaultdict(list)
                for e in items:
                    by_cat[e["category"]].append(e)
                for cat, cat_items in by_cat.items():
                    print(f"\n  [{cat}] {len(cat_items)}건:")
                    for e in cat_items[:15]:  # 카테고리당 최대 15건
                        print(f"    {e['sheet']}:{e['cell']} — {e['message']}")
                    if len(cat_items) > 15:
                        print(f"    ... 외 {len(cat_items) - 15}건")
    else:
        print(f"\n✅ 에러 0건, 경고 0건")


def main():
    """CLI: python recalc.py <xlsx_file> [--json]"""
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <xlsx_file> [--json]")
        sys.exit(1)

    filepath = sys.argv[1]
    json_mode = "--json" in sys.argv

    result = audit_workbook(filepath)

    if json_mode:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        _print_result(result)

    # 에러 존재시 exit code 1
    if result.get("errors"):
        error_count = sum(1 for e in result["errors"] if e["severity"] == "error")
        sys.exit(1 if error_count > 0 else 0)


if __name__ == "__main__":
    main()
